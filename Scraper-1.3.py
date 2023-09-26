import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
from tqdm import tqdm
import os

# Kaufland scraping logic
def scrape_kaufland_product_data(ean):
    url = f"https://www.kaufland.de/item/search/?search_value={ean}"
    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')

        delivery_time_element = soup.find('span', class_='product-delivery-times')
        title_element = soup.find('div', class_='product__title')
        review_numbers_element = soup.find('span', class_='product-rating__count')
        uv_price_element = soup.find('p', class_='price__note--rrp')
        sale_element = soup.find('p', class_='price__note--savings')
        price_element = soup.find('div', class_='price')

        delivery_time = delivery_time_element.get_text(strip=True) if delivery_time_element else None
        title = title_element.get_text(strip=True) if title_element else None
        review_numbers = int(review_numbers_element.get_text(strip=True)) if review_numbers_element else None
        sale = sale_element.get_text(strip=True) if sale_element else None

        price = None
        uv_price = None

        if price_element:
            price_text = price_element.get_text(strip=True)
            price = float(price_text.replace('€', '').replace(',', '.'))

        if uv_price_element:
            uv_price = uv_price_element.get_text(strip=True)
            uv_price = uv_price.replace('UVP', '')
            uv_price = float(uv_price.replace('€', '').replace(',', '.'))

        product_url = url  # Save the product URL

        return {
            'EAN': ean,
            'URL': product_url,
            'Cena': price,
            'UVP': uv_price,
            'Tytuł': title,
            'Czas Wysyłki': delivery_time,
            'Ilość Opinii': review_numbers,
            'Promocja': sale
        }
    else:
        return None

# eBay scraping logic
def scrape_ebay_product_data(ean, country_code):
    country_urls = {
        'de': f"https://www.ebay.de/sch/i.html?_from=R40&_trksid=p2334524.m570.l1313&_nkw={ean}",
        'fr': f"https://www.ebay.fr/sch/i.html?_from=R40&_trksid=p2334524.m570.l1313&_nkw={ean}",
        'it': f"https://www.ebay.it/sch/i.html?_from=R40&_trksid=p2334524.m570.l1313&_nkw={ean}"
    }

    url = country_urls.get(country_code)

    if url is None:
        print(f"Invalid country code: {country_code}")
        return []

    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.content, "html.parser")

        data = []

        product_elements = soup.find_all("div", class_="s-item__info clearfix")

        for product_element in product_elements:
            title_element = product_element.find("div", class_="s-item__title")
            price_element = product_element.find("span", class_="s-item__price")
            shipping_element = product_element.find("span", class_="s-item__shipping s-item__logisticsCost")
            review_element = product_element.find("div", class_="x-star-rating")
            review_count_element = product_element.find("span", class_="s-item__reviews-count")
            delivery_from_element = product_element.find("span", class_="s-item__location s-item__itemLocation")

            title_text = title_element.get_text(strip=True) if title_element else ""
            price_text = price_element.get_text(strip=True) if price_element else ""
            shipping_text = shipping_element.get_text(strip=True) if shipping_element else ""
            review_text = review_element.find("span", class_="clipped").get_text(strip=True) if review_element else ""
            delivery_from_text = delivery_from_element.get_text(strip=True) if delivery_from_element else ""

            if review_count_element:
                review_count_text = review_count_element.find("span", attrs={"aria-hidden": "false"}).get_text(
                    strip=True)
            else:
                review_count_text = ""

            data.append({"EAN": ean, "URL": url, "Cena": price_text, "Tytuł": title_text, "Wysyłka": shipping_text,
                         "Ilość Opinii": review_count_text, "Waga Opinii": review_text,
                         "Kraj Wysyłki": delivery_from_text})

        return data
    else:
        return []

def select_input_file():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(title="Select Input File", filetypes=[("Excel Files", "*.xlsx")])
    return file_path


def main():
    print("Select a marketplace:")
    print("1. Kaufland")
    print("2. eBay")

    choice = input("Enter the number of the marketplace: ")

    if choice == '1':
        marketplace = "Kaufland"
        scrape_product_data = scrape_kaufland_product_data
        headers = ['EAN', 'URL', 'Cena', 'UVP', 'Tytuł', 'Czas Wysyłki', 'Ilość Opinii', 'Promocja', 'Data Analizy']

        # Create a counter for Kaufland files
        file_counter = 1
        while True:
            output_file = f"Kaufland_scraped_data_{file_counter}.xlsx"
            if not os.path.isfile(output_file):
                break
            file_counter += 1

    elif choice == '2':
        marketplace = "eBay"
        print("Select country:")
        print("1. Germany (de)")
        print("2. France (fr)")
        print("3. Italy (it)")
        country_choice = input("Enter the number of the country: ")

        country_codes = {
            '1': 'de',
            '2': 'fr',
            '3': 'it'
        }

        country_code = country_codes.get(country_choice)

        if country_code is None:
            print("Invalid country choice.")
            return

        # Create a counter for eBay files
        file_counter = 1
        while True:
            output_file = f"Ebay.{country_code}_scraped_data_{file_counter}.xlsx"
            if not os.path.isfile(output_file):
                break
            file_counter += 1

        scrape_product_data = lambda ean: scrape_ebay_product_data(ean, country_code)
        headers = ['EAN', 'URL', 'Cena', 'Tytuł', 'Wysyłka', 'Ilość Opinii', 'Waga Opinii', 'Kraj Wysyłki',
                   'Data Analizy']
    else:
        print("Invalid choice.")
        return

    input_file = select_input_file()  # Get the selected input file path

    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    row_number = 2  # Start from the second row

    print("Please wait... Loading...")  # Add this line to show "Pending" message

    # Track processed EANs and the number of sets processed for each EAN
    processed_eans = {}

    # Count the total number of EANs to process
    total_eans = sheet.max_row - 1  # Subtract 1 for the header row

    with tqdm(total=total_eans, desc="Processing EANs") as pbar:
        while row_number <= sheet.max_row:
            ean = sheet.cell(row=row_number, column=1).value

            if ean not in processed_eans:
                processed_eans[ean] = 0  # Initialize the count

            if processed_eans[ean] >= 2:
                row_number += 1
                continue  # Skip if two sets have been processed for this EAN

            product_data = scrape_product_data(ean)

            if product_data:
                timestamp = datetime.now()
                timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')

                if marketplace == "Kaufland":
                    data_dict = product_data
                    data_dict['Data Analizy'] = timestamp_str
                    for col_num, key in enumerate(headers, start=1):
                        if key in data_dict:
                            sheet.cell(row=row_number, column=col_num, value=data_dict[key])
                elif marketplace == "eBay":
                    if len(product_data) >= 2:
                        second_set_data_dict = product_data[1]
                        second_set_data_dict['Data Analizy'] = timestamp_str
                        sheet.cell(row=row_number, column=2, value=second_set_data_dict['URL'])
                        col_num = 3
                        for key in headers[2:]:
                            if key in second_set_data_dict:
                                sheet.cell(row=row_number, column=col_num, value=second_set_data_dict[key])
                                col_num += 1

                processed_eans[ean] += 1
                row_number += 1  # Increment row_number only when valid data is found

            pbar.update(1)  # Update the progress bar


    # Save the file with a unique name
    wb.save(output_file)
    print(f"Scraped data saved to {output_file}")

    print("DONE")  # Print "DONE"
    input("Press Enter to exit...")  # Wait for user input before closing


if __name__ == "__main__":
    main()